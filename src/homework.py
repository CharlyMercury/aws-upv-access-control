import openpyxl


class Homework():

    def __init__(self, students, tareas):
        self.wb = openpyxl.Workbook()
        self.students = students
        self.students_by_tarea = {}
        self.homework_sheet = None
        self.create_tareas(tareas)

    def create_tareas(self, tareas_in):
        for key, value in tareas_in.items():
            self.wb.create_sheet(key)
            self.homework_sheet = self.wb[key]
            self.set_headers(value)
            self.set_students_for_each_homework(key)

    def set_headers(self, tarea):
        self.homework_sheet.cell(row=1, column=1).value = 'Matricula'
        self.homework_sheet.cell(row=1, column=2).value = 'Nombre'
        self.homework_sheet.cell(row=1, column=3).value = tarea
    
    def set_value(self, tarea, row, current_student, name, status):
        self.homework_sheet = self.wb[tarea]
        self.homework_sheet.cell(row=row, column=1).value = current_student
        self.homework_sheet.cell(row=row, column=2).value = name
        self.homework_sheet.cell(row=row, column=3).value = status

    def set_students_for_each_homework(self, tarea):
        self.students_by_tarea[tarea] = []

    def set_student(self, tarea, matricula):
        try: 
            self.students_by_tarea[tarea].append(matricula)
        except: 
            pass

    def save_excel(self, path_excel):
        sheet_default = self.wb['Sheet']
        self.wb.remove(sheet_default)        
        self.wb.save(path_excel)
