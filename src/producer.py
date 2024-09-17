import time
import datetime
import os
import boto3
import json
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import PatternFill



upv_termination = "@upv.edu.mx"
gmail_termination = "@gmail.com"
course = " Metodología de la programación "
attenace_path = "C:\\Users\\carlo\\OneDrive\\upv_clases\\Metodologia_de_la_programacion\\Grupo-IM-20-"
groups = ["75004", "75011", "75018"]
parameters_group = []


for group in groups:

    group_file = f"groups/IM_20-{group}.json"
        
    with open(file=group_file, mode="r", encoding="utf-8") as file:
        parameters_group.append(json.load(file))


def get_group(matricula_):
    try: 
        counter = 0
        for grupo in parameters_group:
            if int(matricula_) in grupo['students_id']:
                grupo = grupo['group']
                break
            else:
                grupo = 0
                counter += 1            
        if counter == len(groups):
            print("No estás en Lista")
        counter = 0
    except Exception as err:
        print(f"Hubo un error: {err}. Funcion get_group")
    return grupo          


def sheet_letter(group_in_2):

    schedule_ = f"groups/horario.json"  
    with open(file=schedule_, mode="r", encoding="utf-8") as file:
        schedule_days = json.load(file)

    timestamp = time.time()
    today_day = datetime.datetime.fromtimestamp(float(timestamp)).strftime('%A').lower()[0:2]
    letter_out = schedule_days[group_in_2][today_day]

    return letter_out


def set_attendance_in_excel(grupo_in, matricula):

    try: 
        workbook = openpyxl.load_workbook(attenace_path+str(grupo_in)+f'\\Asistencia IM 20-{str(grupo_in)}'+'.xlsx')
        sheet = workbook['Semana3']

        for j in range(3,40):
            c3=sheet['C'+str(j)].value
            if c3 == int(matricula):
                break
        
        letter_excel = sheet_letter(str(grupo_in))

        sheet[letter_excel+str(j)].fill = PatternFill(bgColor="4ea72e", fill_type = "solid")
        os.remove(attenace_path+str(grupo_in)+f'\\Asistencia IM 20-{str(grupo_in)}'+'.xlsx')
        time.sleep(1)
        workbook.save(attenace_path+str(grupo_in)+f'\\Asistencia IM 20-{str(grupo_in)}'+'.xlsx')
        workbook.close()
        print("Save Successfully")
    except Exception as err:
        print(f"Hubo un error: {err}. Funcion set_attendance_in_excel")



def producer_student_attendance(parameters, parameters_class):

    """client = boto3.client("sqs",
                        region_name=parameters["aws_region"],
                        aws_access_key_id=parameters["aws_key_id"],
                        aws_secret_access_key=parameters["aws_secret_key"])
    queue_url = parameters["queue_url"]"""

    subject = f"{parameters_class["subject"]}: {course}. "
    day_objective = parameters_class["content"]

    while True:

        matricula = input(" Capturar matrícula: ")

        if matricula == "test":
            matricula = "carlos.mercury92"
            email_termination = gmail_termination
        else:
            email_termination = upv_termination

        if matricula == '1030034':
            break

        grupo = get_group(matricula)
        if str(grupo) in groups:
            set_attendance_in_excel(grupo, matricula)

        timestamp = time.time()

        """response = client.send_message(
            QueueUrl=queue_url,
            MessageBody=json.dumps({
                "matricula": matricula,
                "timestamp": timestamp,
                "subject": subject,
                "content": day_objective,
                "receiver": matricula+email_termination,
            })
        )

        if response["ResponseMetadata"]["HTTPStatusCode"] == 200:
            print(f"\tAlumno {matricula} capturado correctamente\t\n")"""


def producer_student_homework(parameters, parameters_homework):

    client = boto3.client("sqs",
                        region_name=parameters["aws_region"],
                        aws_access_key_id=parameters["aws_key_id"],
                        aws_secret_access_key=parameters["aws_secret_key"])
    queue_url = parameters["queue_url"]

    subject = f" {parameters_homework["subject"]}: {course}. "
    week_homework = parameters_homework["content"]

    group_id = input(" Capturar id grupo (00, 04, 11, 18): ")
    group = f"groups/IM_20-750{group_id}.json"

    if group_id == "00":
        email_termination = gmail_termination
    else:
        email_termination = upv_termination 

    with open(file=group, mode="r", encoding="utf-8") as file:
        parameters_group = json.load(file)

    for student in parameters_group["students_id"]:
        matricula = str(student)
        timestamp = time.time()

        response = client.send_message(
            QueueUrl=queue_url,
            MessageBody=json.dumps({
                "matricula": matricula,
                "timestamp": timestamp,
                "subject": subject,
                "content": week_homework,
                "receiver": matricula+email_termination,
            })
        )

        if response["ResponseMetadata"]["HTTPStatusCode"] == 200:
            print(f"\tAlumno {matricula}: Tarea enviada correctamente. \t\n")


def producer_student_email_notifications(parameters, parameters_homework):
    
    client = boto3.client("sqs",
                        region_name=parameters["aws_region"],
                        aws_access_key_id=parameters["aws_key_id"],
                        aws_secret_access_key=parameters["aws_secret_key"])
    queue_url = parameters["queue_url"]

    subject = f" {parameters_homework["subject"]}: {course}. "
    week_homework = parameters_homework["content"]

    group_id = input(" Capturar id grupo (00, 04, 11, 18): ")
    group = f"groups/IM_20-750{group_id}.json"    

    if group_id == "00":
        email_termination = gmail_termination
    else:
        email_termination = upv_termination  

    with open(file=group, mode="r", encoding="utf-8") as file:
        parameters_group = json.load(file)

    for student in parameters_group["students_id"]:
        matricula = str(student)
        timestamp = time.time()

        response = client.send_message(
            QueueUrl=queue_url,
            MessageBody=json.dumps({
                "matricula": matricula,
                "timestamp": timestamp,
                "subject": subject,
                "content": week_homework,
                "receiver": matricula+email_termination,
            })
        )

        if response["ResponseMetadata"]["HTTPStatusCode"] == 200:
            print(f"\t Aviso enviado correctamente a {matricula}. \t\n")